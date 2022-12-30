from openpyxl import Workbook, load_workbook

wb = load_workbook("pythonAutomation/automateExcel/example.xlsx")

ws = wb.active
print(ws) # <Worksheet "Sheet1">
print(ws['A1'].value) # <Worksheet "Sheet1">
#2015-04-05 13:34:02
# To change the name
ws["A1"].value = "Test"
wb.save("pythonAutomation/automateExcel/example.xlsx")

print(wb.sheetnames) # ['Sheet1', 'Sheet2', 'Sheet3']
# to access sheet
print(wb["Sheet1"])
# to create sheet
wb.create_sheet("Test")

print(wb.sheetnames) 
# if you want to create sheet in actual sheet you have to save it wb.save("pythonAutomation/automateExcel/example.xlsx")

