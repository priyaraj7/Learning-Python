# from openpyxl import Workbook, load_workbook

# wb = Workbook() # create workbook
# ws = wb.active
# ws.title = "Data"

# ws.append(["Name", "Favorite", "City"]) # append rows

# wb.save("pythonAutomation/automateExcel/test.xlsx")

# ACCESSING MULTIPLE CELLS

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

wb = load_workbook("pythonAutomation/automateExcel/test.xlsx")
ws = wb.active

for row in range(1, 11):
    for col in range(1,5):
        char = get_column_letter(col)
        ws[char +str(row)] = char + str(row)

# # MERGING CELL
# ws.merged_cells("A1:D1") # once you merge you are going to loose the date
# ws.unmerged_cells("A1:D1") # once you merge you are going to loose the date

# # INSERT / DELETE ROWS
# ws.insert_rows(7)
# # ws.insert_rows(9)
# ws.delete_rows(9)

wb.save("pythonAutomation/automateExcel/test.xlsx")